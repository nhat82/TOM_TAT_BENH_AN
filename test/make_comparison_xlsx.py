"""
Build an Excel comparison from two benchmark_models.py JSON reports.

Usage:
  python test/make_comparison_xlsx.py
  python test/make_comparison_xlsx.py report_a.json report_b.json --out out.xlsx
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BETTER_FILL = PatternFill("solid", fgColor="C6EFCE")
WORSE_FILL = PatternFill("solid", fgColor="FFC7CE")
BOLD = Font(bold=True)


def _load(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def _style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build(report_a: dict, report_b: dict, out_path: Path) -> None:
    wb = Workbook()

    # ── sheet 1: summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    model_a, model_b = report_a["model"], report_b["model"]
    sa, sb = report_a["summary"], report_b["summary"]

    headers = ["Metric", model_a, model_b, "Better"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    rows = [
        ("TTFT avg (s)", sa["ttft_avg_s"], sb["ttft_avg_s"], "lower"),
        ("TTFT p50 (s)", sa["ttft_p50_s"], sb["ttft_p50_s"], "lower"),
        ("Error rate", sa["error_rate"], sb["error_rate"], "lower"),
        ("Tool calls avg", sa["tool_calls_avg"], sb["tool_calls_avg"], "lower"),
    ]
    for name, va, vb, direction in rows:
        r = ws.max_row + 1
        winner = model_a if (va is not None and vb is not None and (va < vb) == (direction == "lower")) else model_b
        if va is None or vb is None or va == vb:
            winner = "-"
        ws.append([name, va, vb, winner])
        if va is not None and vb is not None and va != vb:
            better_col, worse_col = (2, 3) if (va < vb) == (direction == "lower") else (3, 2)
            ws.cell(row=r, column=better_col).fill = BETTER_FILL
            ws.cell(row=r, column=worse_col).fill = WORSE_FILL

    for c in range(1, 5):
        ws.cell(row=1, column=c)
    ws.cell(row=1, column=1).font = HEADER_FONT
    _autosize(ws, {1: 20, 2: 24, 3: 24, 4: 24})

    # ── sheet 2: per-question ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Per-Question")
    headers2 = [
        "Patient ID", "Question",
        f"TTFT {model_a} (s)", f"TTFT {model_b} (s)",
        f"Tool calls {model_a}", f"Tool calls {model_b}",
        f"Error {model_a}", f"Error {model_b}",
    ]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))

    qa_by_key = {(q["patient_id"], q["question"]): q for q in report_a["per_question"]}
    qb_by_key = {(q["patient_id"], q["question"]): q for q in report_b["per_question"]}
    keys = list(qa_by_key.keys())  # preserve order from report A

    for key in keys:
        qa = qa_by_key.get(key, {})
        qb = qb_by_key.get(key, {})
        pid, question = key
        ws2.append([
            pid, question,
            qa.get("ttft_s"), qb.get("ttft_s"),
            qa.get("tool_calls"), qb.get("tool_calls"),
            "ERROR" if qa.get("error") else "",
            "ERROR" if qb.get("error") else "",
        ])
        r = ws2.max_row
        ta, tb = qa.get("ttft_s"), qb.get("ttft_s")
        if ta is not None and tb is not None and ta != tb:
            better_col, worse_col = (3, 4) if ta < tb else (4, 3)
            ws2.cell(row=r, column=better_col).fill = BETTER_FILL
            ws2.cell(row=r, column=worse_col).fill = WORSE_FILL
        if qa.get("error"):
            ws2.cell(row=r, column=7).fill = WORSE_FILL
        if qb.get("error"):
            ws2.cell(row=r, column=8).fill = WORSE_FILL

    _autosize(ws2, {1: 14, 2: 55, 3: 16, 4: 16, 5: 14, 6: 14, 7: 30, 8: 30})
    ws2.freeze_panes = "A2"

    # ── sheet 3: errors ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Errors")
    ws3.append(["Model", "Patient ID", "Question", "Error"])
    _style_header(ws3, 1, 4)
    for report in (report_a, report_b):
        for q in report["per_question"]:
            if q.get("error"):
                ws3.append([report["model"], q["patient_id"], q["question"], q["error"]])
    _autosize(ws3, {1: 24, 2: 14, 3: 50, 4: 70})

    wb.save(out_path)
    print(f"Excel report -> {out_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build an Excel comparison from two benchmark reports.")
    parser.add_argument("report_a", nargs="?", default="test/benchmark_report_local_qwen3_14b.json")
    parser.add_argument("report_b", nargs="?", default="test/benchmark_report_api_gemini-3.1-flash-lite.json")
    parser.add_argument("--out", default="test/model_comparison.xlsx")
    args = parser.parse_args()

    report_a = _load(Path(args.report_a))
    report_b = _load(Path(args.report_b))
    build(report_a, report_b, Path(args.out))


if __name__ == "__main__":
    main()
